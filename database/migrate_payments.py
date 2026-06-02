#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Diamond Cleaning - Payment Data Migration
Reads payments Excel and generates SQL INSERT statements for contracts and payments
"""

import openpyxl
import re
import sys
import os
from datetime import datetime, timedelta

# ============================================================
# Configuration
# ============================================================
PAYMENTS_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    '..', '\u0627\u0644\u062a\u062d\u0635\u064a\u0644\u0627\u062a (2) (1).xlsx')
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '06_payment_data.sql')

# Receiver name mapping to receiver IDs (from seed data)
RECEIVER_MAP = {
    '\u0639\u0628\u062f\u0627\u0644\u0644\u0647': 1,
    '\u062d\u0628\u064a\u0628': 2,
    '\u0631\u0636\u0648\u0627\u0646': 3,
    '\u0639\u0627\u0644\u0645': 4,
    '\u0643\u0627\u0634': 5,
    '\u062d\u0648\u0627\u0644\u0629 \u0627\u0644\u0634\u0631\u0643\u0629': 6,
    '\u062d\u0648\u0627\u0644\u0647 \u0627\u0644\u0634\u0631\u0643\u0647': 6,
    '\u062d\u0648\u0627\u0644\u0629 \u0628\u0646\u0643\u064a\u0629 \u0627\u0644\u0623\u0647\u0644\u064a': 7,
    '\u062d\u0648\u0627\u0644\u0647 \u0628\u0646\u0643\u064a\u0647 \u0627\u0644\u0627\u0647\u0644\u064a': 7,
}

# Payment period to months mapping
PERIOD_TO_MONTHS = {
    '1 \u0634\u0647\u0631': 1,
    '\u0634\u0647\u0631': 1,
    '\u0634\u0647\u0631\u064a': 1,
    '2 \u0627\u0634\u0647\u0631': 2,
    '3 \u0627\u0634\u0647\u0631': 3,
    '4 \u0627\u0634\u0647\u0631': 4,
    '6 \u0627\u0634\u0647\u0631': 6,
    '\u0633\u0646\u0647': 12,
    '\u0633\u0646\u0629': 12,
}


def clean_phone(phone_val):
    """Clean and normalize phone number"""
    if phone_val is None:
        return None
    phone = str(phone_val).strip()
    if phone.endswith('.0'):
        phone = phone[:-2]
    phone = re.sub(r'[\s\-\(\)\+]', '', phone)
    if phone.startswith('966') and len(phone) == 12:
        return phone
    if phone.startswith('0') and len(phone) == 10:
        return '966' + phone[1:]
    if len(phone) == 9 and phone[0] == '5':
        return '966' + phone
    return phone if phone else None


def escape_sql(val):
    """Escape string for SQL"""
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if s == '' or s == 'None' or s == 'nan':
        return 'NULL'
    s = s.replace("'", "''")
    return f"E'{s}'"


def main():
    # Try to find payments file
    payments_file = PAYMENTS_FILE
    if not os.path.exists(payments_file):
        # Try current directory
        alt_paths = [
            os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '..', '\u0627\u0644\u062a\u062d\u0635\u064a\u0644\u0627\u062a (2) (1).xlsx'),
        ]
        for alt in alt_paths:
            if os.path.exists(alt):
                payments_file = alt
                break
        else:
            print(f"ERROR: Cannot find payments file")
            print(f"Tried: {PAYMENTS_FILE}")
            sys.exit(1)

    print(f"Reading: {payments_file}")

    wb = openpyxl.load_workbook(payments_file, read_only=True)
    ws = wb[wb.sheetnames[0]]  # الورقة1

    # Structure:
    # Row 1: Headers (cols 0-8) + year markers (col 9=2025, col 21=2026, col 33=2027)
    # Row 2: Month numbers (1-12 under each year)
    # Row 3+: Customer data

    # Column mapping:
    # 0: code, 1: customer name, 2: phone, 3: contract_start
    # 4: status, 5: payment_method (مقدم/متأخر), 6: period, 7: amount
    # 8: receiver
    # 9-20: 2025 months (1-12)
    # 21-32: 2026 months (1-12)
    # 33-44: 2027 months (1-12)

    payment_customers = []
    all_payments = []
    contracts_data = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        vals = list(row)
        if vals[0] is None or vals[1] is None:
            continue

        customer_name = str(vals[1]).strip()
        phone = clean_phone(vals[2])
        contract_start = vals[3]  # datetime or None
        status = str(vals[4]).strip() if vals[4] else '\u064a\u0639\u0645\u0644'
        payment_method = str(vals[5]).strip() if vals[5] else None  # مقدم/متأخر
        period_str = str(vals[6]).strip() if vals[6] else None
        total_amount = float(vals[7]) if vals[7] else None
        receiver_name = str(vals[8]).strip() if vals[8] else None

        # Get receiver ID
        receiver_id = None
        if receiver_name:
            receiver_id = RECEIVER_MAP.get(receiver_name)
            if receiver_id is None:
                # Try fuzzy matching
                for key, rid in RECEIVER_MAP.items():
                    if key in receiver_name or receiver_name in key:
                        receiver_id = rid
                        break

        # Calculate contract duration
        duration_months = PERIOD_TO_MONTHS.get(period_str, 1) if period_str else None

        # Calculate end date
        contract_start_date = None
        contract_end_date = None
        if contract_start and isinstance(contract_start, datetime):
            contract_start_date = contract_start.strftime('%Y-%m-%d')
            if duration_months:
                end = contract_start
                # Add months
                month = contract_start.month + duration_months
                year = contract_start.year + (month - 1) // 12
                month = (month - 1) % 12 + 1
                try:
                    contract_end_date = datetime(year, month, min(contract_start.day, 28)).strftime('%Y-%m-%d')
                except:
                    contract_end_date = None

        # Monthly subscription
        monthly_sub = None
        if total_amount and duration_months and duration_months > 0:
            monthly_sub = round(total_amount / duration_months, 2)

        # Store contract data
        contracts_data.append({
            'customer_name': customer_name,
            'phone': phone,
            'start_date': contract_start_date,
            'end_date': contract_end_date,
            'duration_months': duration_months,
            'total_amount': total_amount,
            'monthly_subscription': monthly_sub,
            'payment_type': payment_method,
            'payment_period': period_str,
            'receiver_name': receiver_name,
            'receiver_id': receiver_id,
        })

        # Extract monthly payments
        # 2025: cols 9-20 (months 1-12)
        for m in range(12):
            col = 9 + m
            if col < len(vals) and vals[col] is not None:
                try:
                    amount = float(vals[col])
                    if amount > 0:
                        all_payments.append({
                            'customer_name': customer_name,
                            'phone': phone,
                            'amount': amount,
                            'year': 2025,
                            'month': m + 1,
                            'payment_date': f'2025-{m+1:02d}-01',
                            'receiver_id': receiver_id,
                            'receiver_name': receiver_name,
                            'payment_method': receiver_name,
                        })
                except (ValueError, TypeError):
                    pass

        # 2026: cols 21-32
        for m in range(12):
            col = 21 + m
            if col < len(vals) and vals[col] is not None:
                try:
                    amount = float(vals[col])
                    if amount > 0:
                        all_payments.append({
                            'customer_name': customer_name,
                            'phone': phone,
                            'amount': amount,
                            'year': 2026,
                            'month': m + 1,
                            'payment_date': f'2026-{m+1:02d}-01',
                            'receiver_id': receiver_id,
                            'receiver_name': receiver_name,
                            'payment_method': receiver_name,
                        })
                except (ValueError, TypeError):
                    pass

        # 2027: cols 33-44
        for m in range(12):
            col = 33 + m
            if col < len(vals) and vals[col] is not None:
                try:
                    amount = float(vals[col])
                    if amount > 0:
                        all_payments.append({
                            'customer_name': customer_name,
                            'phone': phone,
                            'amount': amount,
                            'year': 2027,
                            'month': m + 1,
                            'payment_date': f'2027-{m+1:02d}-01',
                            'receiver_id': receiver_id,
                            'receiver_name': receiver_name,
                            'payment_method': receiver_name,
                        })
                except (ValueError, TypeError):
                    pass

    wb.close()

    # Generate SQL
    lines = []
    lines.append('-- ============================================================')
    lines.append('-- Diamond Cleaning - Payment & Contract Data (Auto-Generated)')
    lines.append(f'-- Contracts: {len(contracts_data)}')
    lines.append(f'-- Payments: {len(all_payments)}')
    lines.append('-- ============================================================')
    lines.append('')

    # Insert contracts
    lines.append('-- ============================================================')
    lines.append('-- Contracts')
    lines.append('-- ============================================================')
    lines.append('')

    for i, ct in enumerate(contracts_data):
        if ct['total_amount'] is None and ct['start_date'] is None:
            continue  # Skip contracts with no useful data

        start_date = f"'{ct['start_date']}'" if ct['start_date'] else 'NULL'
        end_date = f"'{ct['end_date']}'" if ct['end_date'] else 'NULL'
        duration = ct['duration_months'] if ct['duration_months'] else 'NULL'
        total = ct['total_amount'] if ct['total_amount'] else 'NULL'
        monthly = ct['monthly_subscription'] if ct['monthly_subscription'] else 'NULL'

        # Find customer by name (using subquery)
        customer_name_escaped = ct['customer_name'].replace("'", "''")

        lines.append(f"INSERT INTO contracts (customer_id, start_date, end_date, duration_months, total_amount, monthly_subscription, payment_type, payment_period, status)")
        lines.append(f"SELECT c.id, {start_date}, {end_date}, {duration}, {total}, {monthly}, {escape_sql(ct['payment_type'])}, {escape_sql(ct['payment_period'])}, E'\\u0646\\u0634\\u0637'")
        lines.append(f"FROM customers c WHERE c.name_ar = E'{customer_name_escaped}' LIMIT 1;")
        lines.append('')

    # Insert payments
    lines.append('')
    lines.append('-- ============================================================')
    lines.append('-- Payments')
    lines.append('-- ============================================================')
    lines.append('')

    for i, p in enumerate(all_payments):
        customer_name_escaped = p['customer_name'].replace("'", "''")
        receiver_id = p['receiver_id'] if p['receiver_id'] else 'NULL'

        lines.append(f"INSERT INTO payments (customer_id, amount, payment_date, receiver_id, payment_method)")
        lines.append(f"SELECT c.id, {p['amount']}, '{p['payment_date']}', {receiver_id}, {escape_sql(p['payment_method'])}")
        lines.append(f"FROM customers c WHERE c.name_ar = E'{customer_name_escaped}' LIMIT 1;")
        lines.append('')

    # Write output
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"\nGenerated: {OUTPUT_FILE}")
    print(f"Total contracts: {sum(1 for c in contracts_data if c['total_amount'] or c['start_date'])}")
    print(f"Total payments: {len(all_payments)}")
    print(f"Total amount: {sum(p['amount'] for p in all_payments):,.0f} SAR")

    # Payment breakdown by year
    yearly = {}
    for p in all_payments:
        yearly[p['year']] = yearly.get(p['year'], 0) + p['amount']
    print(f"\nYearly breakdown:")
    for year in sorted(yearly):
        print(f"  {year}: {yearly[year]:,.0f} SAR")

    # Receiver breakdown
    recv = {}
    for p in all_payments:
        r = p['receiver_name'] or 'Unknown'
        recv[r] = recv.get(r, 0) + p['amount']
    print(f"\nReceiver breakdown:")
    for r, amount in sorted(recv.items(), key=lambda x: -x[1]):
        print(f"  {r}: {amount:,.0f} SAR")


if __name__ == '__main__':
    main()
