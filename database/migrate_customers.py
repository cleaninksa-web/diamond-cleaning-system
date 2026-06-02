#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Diamond Cleaning - Customer Data Migration
Reads customers.xlsx and generates SQL INSERT statements
"""

import openpyxl
import re
import sys
import os

# ============================================================
# Configuration
# ============================================================
EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    '..', 'diamond-cleaning-antigravity-package', 'antigravity_package', 'assets', 'customers.xlsx')
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '05_customer_data.sql')

# Technician name mapping (merge "محمد ساجد" into "ساجد")
TECHNICIAN_MAP = {
    'حبيب': 1,
    'رضوان': 2,
    'ساجد': 3,
    'محمد ساجد': 3,  # Merged into ساجد
    'عالم قير': 4,
}

# District name normalization
DISTRICT_NORMALIZE = {
    'ابحر الشمالية': 'أبحر الشمالية',
    'ابحرالشمالية': 'أبحر الشمالية',
    'ابحر الشماليه': 'أبحر الشمالية',
    'أبحر الشمالية': 'أبحر الشمالية',
    'ابحر الجنوبية': 'أبحر الجنوبية',
    'ابحر الجنوبيه': 'أبحر الجنوبية',
    'ابحرالجنوبية': 'أبحر الجنوبية',
    'المحمدية': 'المحمدية',
    'البساتين': 'البساتين',
    'الرحيلي': 'الرحيلي',
    'الشاطي': 'الشاطئ',
    'الشاطئ': 'الشاطئ',
    'المرجان': 'المرجان',
    'الروضة': 'الروضة',
    'الاصالة': 'الأصالة',
    'الحمراء': 'الحمراء',
    'حراء': 'حراء',
    'الجامعة': 'الجامعة',
}

# Visit count mapping
VISIT_COUNT_MAP = {
    'زيارة': 1,
    'زيارتين': 2,
    '3 زيارات': 3,
    '4 زيارات': 4,
}


def clean_phone(phone_val):
    """Clean and normalize phone number"""
    if phone_val is None:
        return None
    phone = str(phone_val).strip()
    # Remove .0 from float conversion
    if phone.endswith('.0'):
        phone = phone[:-2]
    # Remove spaces, dashes, parentheses
    phone = re.sub(r'[\s\-\(\)\+]', '', phone)
    # If starts with 966, keep as is
    if phone.startswith('966') and len(phone) == 12:
        return phone
    # If starts with 0, prepend 966
    if phone.startswith('0') and len(phone) == 10:
        return '966' + phone[1:]
    # If 9 digits (no prefix), prepend 966
    if len(phone) == 9 and phone[0] == '5':
        return '966' + phone
    return phone if phone else None


def escape_sql(val):
    """Escape string for SQL"""
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if s == '' or s == 'None' or s == 'nan':
        return 'NULL'
    s = s.replace("'", "''")
    return f"E'{s}'"


def parse_visit_days(days_str):
    """Parse visit days string and return day booleans"""
    if not days_str:
        return {'sat': False, 'sun': False, 'mon': False, 'tue': False,
                'wed': False, 'thu': False, 'fri': False}

    days_str = str(days_str).strip().lower()
    result = {
        'sat': any(d in days_str for d in ['السبت', 'سبت']),
        'sun': any(d in days_str for d in ['الأحد', 'الاحد', 'أحد', 'احد']),
        'mon': any(d in days_str for d in ['الإثنين', 'الاثنين', 'إثنين', 'اثنين']),
        'tue': any(d in days_str for d in ['الثلاثاء', 'ثلاثاء']),
        'wed': any(d in days_str for d in ['الأربعاء', 'الاربعاء', 'أربعاء', 'اربعاء']),
        'thu': any(d in days_str for d in ['الخميس', 'خميس']),
        'fri': any(d in days_str for d in ['الجمعة', 'جمعة']),
    }
    return result


def main():
    print(f"Reading: {EXCEL_FILE}")

    if not os.path.exists(EXCEL_FILE):
        # Try alternative path
        alt_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            '..', 'diamond-cleaning-antigravity-package', 'antigravity_package', 'assets', 'customers.xlsx')
        if os.path.exists(alt_path):
            excel_file = alt_path
        else:
            print(f"ERROR: Cannot find {EXCEL_FILE}")
            print("Please ensure customers.xlsx is in the correct location")
            sys.exit(1)
    else:
        excel_file = EXCEL_FILE

    wb = openpyxl.load_workbook(excel_file, read_only=True)
    ws = wb[wb.sheetnames[0]]  # الورقة1

    # Column mapping (0-indexed):
    # 0: ID, 1: name_ar, 2: name_en, 3: phone, 4: driver_phone, 5: maid_phone
    # 6: status, 7: district, 8: pool_count, 9: pool_type
    # 10: weekly_visits, 11: technician, 12: visit_days, 13: monthly_visits
    # 14: notes, 15: location, 16-22: photos, 23: equipment_info
    # 24: lights_count, 25: filter_size, 26: length, 27: width
    # 28: depth_1, 29: depth_2, 30: pool_area, 31: tile_type
    # 32: equipment_notes, 33: contract_value, 34: payment_type
    # 35: payment_nature, 36: contract_start, 37: monthly_fee, 38: free_months

    customers = []
    pool_details = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if vals[1] is None:
            continue

        legacy_id = int(vals[0]) if vals[0] else None
        name_ar = str(vals[1]).strip() if vals[1] else None
        name_en = str(vals[2]).strip() if vals[2] else None

        phone = clean_phone(vals[3])
        driver_phone = clean_phone(vals[4])
        maid_phone = clean_phone(vals[5])

        status_raw = str(vals[6]).strip() if vals[6] else 'يعمل'
        status = status_raw  # Keep as-is: يعمل / متوقف

        district_raw = str(vals[7]).strip() if vals[7] else None
        district = DISTRICT_NORMALIZE.get(district_raw, district_raw) if district_raw else None

        tech_name = str(vals[11]).strip() if vals[11] else None
        tech_id = TECHNICIAN_MAP.get(tech_name) if tech_name else None

        notes = str(vals[14]).strip() if vals[14] else None
        location = str(vals[15]).strip() if vals[15] else None

        # Pool details
        pool_count = int(vals[8]) if vals[8] else 1
        pool_type = str(vals[9]).strip() if vals[9] else None
        # Normalize pool types
        if pool_type:
            pool_type_lower = pool_type.lower()
            if 'سكمير' in pool_type_lower or 'skimmer' in pool_type_lower:
                pool_type = 'سكمير'
            elif 'اوفر' in pool_type_lower or 'أوفر' in pool_type_lower or 'overflow' in pool_type_lower:
                pool_type = 'أوفر فلو'
            elif 'فايبر' in pool_type_lower or 'fiber' in pool_type_lower:
                pool_type = 'فايبر'

        weekly_visits_str = str(vals[10]).strip() if vals[10] else None
        weekly_visits = VISIT_COUNT_MAP.get(weekly_visits_str, 2) if weekly_visits_str else 2

        visit_days_str = str(vals[12]).strip() if vals[12] else None
        visit_day_bools = parse_visit_days(visit_days_str)

        # Equipment info
        equipment_info = str(vals[23]).strip() if len(vals) > 23 and vals[23] else None
        lights_count = int(vals[24]) if len(vals) > 24 and vals[24] else None
        filter_size = str(vals[25]).strip() if len(vals) > 25 and vals[25] else None
        pool_length = float(vals[26]) if len(vals) > 26 and vals[26] else None
        pool_width = float(vals[27]) if len(vals) > 27 and vals[27] else None
        depth_1 = float(vals[28]) if len(vals) > 28 and vals[28] else None
        depth_2 = float(vals[29]) if len(vals) > 29 and vals[29] else None
        pool_area = str(vals[30]).strip() if len(vals) > 30 and vals[30] else None
        tile_type = str(vals[31]).strip() if len(vals) > 31 and vals[31] else None
        equipment_notes = str(vals[32]).strip() if len(vals) > 32 and vals[32] else None

        customers.append({
            'legacy_id': legacy_id,
            'name_ar': name_ar,
            'name_en': name_en,
            'phone': phone,
            'driver_phone': driver_phone,
            'maid_phone': maid_phone,
            'status': status,
            'district': district,
            'technician_id': tech_id,
            'general_notes': notes,
            'location_url': location,
        })

        pool_details.append({
            'pool_type': pool_type,
            'pool_count': pool_count,
            'length': pool_length,
            'width': pool_width,
            'depth_1': depth_1,
            'depth_2': depth_2,
            'total_area': pool_area,
            'tile_type': tile_type,
            'filter_size': filter_size,
            'lights_count': lights_count,
            'equipment_info': equipment_info,
            'equipment_notes': equipment_notes,
            'visit_days': visit_days_str,
            'weekly_visits': weekly_visits,
            'visit_day_bools': visit_day_bools,
        })

    wb.close()

    # Generate SQL
    lines = []
    lines.append('-- ============================================================')
    lines.append('-- Diamond Cleaning - Customer Data (Auto-Generated)')
    lines.append(f'-- Total customers: {len(customers)}')
    lines.append('-- ============================================================')
    lines.append('')

    # Insert customers
    for i, c in enumerate(customers):
        tech_id = c['technician_id'] if c['technician_id'] else 'NULL'
        legacy_id = c['legacy_id'] if c['legacy_id'] else 'NULL'

        lines.append(f"INSERT INTO customers (legacy_id, name_ar, name_en, phone, driver_phone, maid_phone, status, district, technician_id, general_notes, location_url)")
        lines.append(f"VALUES ({legacy_id}, {escape_sql(c['name_ar'])}, {escape_sql(c['name_en'])}, {escape_sql(c['phone'])}, {escape_sql(c['driver_phone'])}, {escape_sql(c['maid_phone'])}, {escape_sql(c['status'])}, {escape_sql(c['district'])}, {tech_id}, {escape_sql(c['general_notes'])}, {escape_sql(c['location_url'])});")
        lines.append('')

    lines.append('')
    lines.append('-- ============================================================')
    lines.append('-- Pool Details')
    lines.append('-- ============================================================')
    lines.append('')

    for i, pd in enumerate(pool_details):
        customer_num = i + 1  # 1-indexed since we insert sequentially

        has_pool_data = any([
            pd['pool_type'], pd['length'], pd['width'],
            pd['depth_1'], pd['depth_2'], pd['tile_type'],
            pd['filter_size'], pd['lights_count'],
            pd['equipment_info'], pd['equipment_notes']
        ])

        if has_pool_data or pd['pool_type']:
            pool_count = pd['pool_count'] if pd['pool_count'] else 1
            length = pd['length'] if pd['length'] else 'NULL'
            width = pd['width'] if pd['width'] else 'NULL'
            depth_1 = pd['depth_1'] if pd['depth_1'] else 'NULL'
            depth_2 = pd['depth_2'] if pd['depth_2'] else 'NULL'
            lights = pd['lights_count'] if pd['lights_count'] else 'NULL'

            lines.append(f"INSERT INTO pool_details (customer_id, pool_type, pool_count, length, width, depth_1, depth_2, total_area, tile_type, filter_size, lights_count, equipment_info, equipment_notes)")
            lines.append(f"VALUES ((SELECT id FROM customers WHERE legacy_id = {i+1}), {escape_sql(pd['pool_type'])}, {pool_count}, {length}, {width}, {depth_1}, {depth_2}, {escape_sql(pd['total_area'])}, {escape_sql(pd['tile_type'])}, {escape_sql(pd['filter_size'])}, {lights}, {escape_sql(pd['equipment_info'])}, {escape_sql(pd['equipment_notes'])});")
            lines.append('')

    # Write output
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"\nGenerated: {OUTPUT_FILE}")
    print(f"Total customers: {len(customers)}")
    print(f"  Active: {sum(1 for c in customers if c['status'] == 'يعمل')}")
    print(f"  Stopped: {sum(1 for c in customers if c['status'] == 'متوقف')}")

    # Stats
    districts = {}
    for c in customers:
        d = c['district'] or 'Unknown'
        districts[d] = districts.get(d, 0) + 1

    print(f"\nDistrict distribution (after normalization):")
    for d, count in sorted(districts.items(), key=lambda x: -x[1])[:10]:
        print(f"  {d}: {count}")

    techs = {1: 0, 2: 0, 3: 0, 4: 0}
    for c in customers:
        if c['technician_id']:
            techs[c['technician_id']] = techs.get(c['technician_id'], 0) + 1
    tech_names = {1: 'حبيب', 2: 'رضوان', 3: 'ساجد', 4: 'عالم قير'}
    print(f"\nTechnician distribution (after merge):")
    for tid, count in techs.items():
        print(f"  {tech_names[tid]}: {count}")


if __name__ == '__main__':
    main()
